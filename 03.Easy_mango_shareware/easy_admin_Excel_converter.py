from msilib.schema import RemoveFile
from posixpath import split
import win32com.client as win32
import openpyxl
from tkinter import filedialog
from PyQt5.QtWidgets import *
from PyQt5 import uic
from cryptography.fernet import Fernet
import datetime as dt
import sys, time, os, getmac, webbrowser


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UI_PATH = "easy_admin_Excel_converter.ui"

class MainDialog(QDialog) :
    def __init__(self) :
        QDialog.__init__(self, None)
        uic.loadUi(os.path.join(BASE_DIR, UI_PATH), self)

        self.make_admin_file_btn.setEnabled(False)
        self.make_mangofile_btn.setEnabled(False)

        self.decryptography()

        self.excel_converter_manual_btn.setText('이지어드민-더망고\n\n엑셀변환기 메뉴얼')

#-------------------------------------이지어드민_파일_생성_버튼_관련함수------------------------------------------------------------------------
        # 더망고_파일_선택_파일찾기_버튼
        self.search_mango_xls_btn.clicked.connect(self.search_mango_xls_start)
        # 이지어드민_파일_생성_버튼
        self.make_admin_file_btn.clicked.connect(self.make_adminfile)
#-------------------------------------더망고_송장_파일_생성_버튼_관련함수------------------------------------------------------------------------
        # 이지어드민_파일_선택_파일찾기_버튼
        self.search_admin_xls_btn.clicked.connect(self.search_invoice_admin_xls_start)
        # 망고_파일_선택_파일찾기_버튼
        self.search_mango_xlsx_btn.clicked.connect(self.search_maked_mango_xlsx_start)
        # 더망고_송장_파일_생성_버튼
        self.make_mangofile_btn.clicked.connect(self.make_invoice_mangofile_btn)
#--------------------------------------라이센스 버튼 관련 함수-----------------------------------------------------------------------------------------
        # 라이센스 키 입력/수정 버튼(라이센스 입력 함수 실행)
        self.input_license_btn.clicked.connect(self.input_license)
        # 라이센스 구매링크로 이동
        self.smartstore_btn.clicked.connect(self.smartstore)
        # 엑셀변환 사용법 블로그 링크로 이동
        self.excel_converter_manual_btn.clicked.connect(self.excel_converter_manual_blog)

#-------------------------------------이지어드민_파일_생성_버튼_관련함수------------------------------------------------------------------------
    # 더망고_파일_선택_파일찾기_버튼_함수
    def search_mango_xls_start(self) :
        # 더망고 xls 파일 경로 선택
        mango_xls_fname = filedialog.askopenfilename(initialdir="/", title = "더망고에서 다운받은 xlsx파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.mango_xls_box.setText(mango_xls_fname)
        
    # 이지어드민_파일_생성_버튼_함수
    def make_adminfile(self) :
        # 파일에 입력한 문자열 가져오기
        mango_xls_fname = self.mango_xls_box.toPlainText()

        # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
        mango_xls_fname = mango_xls_fname.replace('/', '\\')

        # 더망고 xls 파일을 선택하지 않았을 시 경고 메시지.
        if mango_xls_fname == "" :
            self.status_label.setText("더망고에서 다운받은 xls 파일을 선택해주세요.")
            return 0
        
        else :
            self.status_label.setText("easy admin 업로드용 파일 생성이 진행중입니다...")
            QApplication.processEvents()

            # xls to xlsx convert
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xls_wb = excel.Workbooks.Open(mango_xls_fname)
            mango_xls_wb.SaveAs(f'{mango_xls_fname}x', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xls_wb.Close()
            excel.Application.Quit()

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 파일 날짜명
            admin_date_file_name_list = mango_xls_fname.split('.')
            admin_date_file_name = admin_date_file_name_list[0]

            # 더망고 xlsx 파일 불러오기
            mango_wb = openpyxl.load_workbook(f'{mango_xls_fname}x', data_only=True)

            # 더망고 xlsx 파일 활성화된 시트 선택
            mango_ws = mango_wb.active

            # 이지어드민 엑셀 파일 생성
            admin_wb = openpyxl.Workbook()
            admin_ws = admin_wb.active


            # 첫 행 양식 생성 방법
            data = []
            for column in range(1, mango_ws.max_column + 1) :
                data.append(mango_ws.cell(1,column).value)
            admin_ws.append(data)

            # 데이터 이전
            for row in mango_ws.iter_rows(min_row=2) :
                if row[10].value == None :
                    admin_ws.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value,                        row[10].value, row[11].value, row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value, row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value, row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value, row[36].value])
                else :
                    admin_ws.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, f'{row[9].value} {row[10].value}', None,            row[11].value, row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value, row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value, row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value, row[36].value])

            # 이지어드민 엑셀 저장
            admin_wb.save(f'{admin_date_file_name}_변환.xlsx')

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # xls to xlsx convert
            excel2 = win32.gencache.EnsureDispatch('Excel.Application')
            admin_xlsx_wb = excel2.Workbooks.Open(f'{admin_date_file_name}_변환.xlsx')
            admin_xlsx_wb.SaveAs(f'{admin_date_file_name}_이지어드민_업로드용.xlsx', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            admin_xlsx_wb.Close()
            excel2.Application.Quit()

            # 위치 자동 내려받기
            self.mango_xlsx_box.setText(f'{mango_xls_fname}x')
            
            self.status_label.setText("이지어드민 업로드용 파일 생성이 완료되었습니다.")
#-------------------------------------더망고_송장_파일_생성_버튼_함수------------------------------------------------------------------------
    # 이지어드민_파일_선택_파일찾기_버튼_함수
    def search_invoice_admin_xls_start(self) : 
        # 이지어드민 송장 xls 파일 경로 선택
        invoice_admin_xls_fname = filedialog.askopenfilename(initialdir="/", title = "CN Plus에서 다운받은 송장 xls파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.admin_xls_box.setText(invoice_admin_xls_fname)
    # 더망고_파일_선택_파일찾기_버튼_함수
    def search_maked_mango_xlsx_start(self) : 
        # 더망고 xlsx 생성 파일 경로 선택
        maked_mango_xlsx_fname = filedialog.askopenfilename(initialdir="/", title = "생성된 더망고 xlsx파일을 선택 해 주세요", filetypes = (("xlsx file","*.xlsx"),("all file","*.*")))
        self.mango_xlsx_box.setText(maked_mango_xlsx_fname)
    # 더망고_송장_파일_생성_버튼_함수
    def make_invoice_mangofile_btn(self) :
        self.status_label2.setText("더망고 송장 파일 생성 중입니다.....")
        QApplication.processEvents()
        # 파일에 입력한 문자열 가져오기
        admin_xls_fpath = self.admin_xls_box.toPlainText()
        mango_xlsx_fpath = self.mango_xlsx_box.toPlainText()

        # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
        admin_xls_fpath = admin_xls_fpath.replace('/', '\\')

        if admin_xls_fpath == "" or mango_xlsx_fpath == "" :
            self.status_label2.setText("이지어드민, 더망고 파일 위치를 선택해주세요.")
            return 0

        else :
            # xls to xlsx convert      # xls 파일이 이상해서 읽어지지 않음. 그래서 xlsx 파일로 변환.
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            admin_xls_converter_wb = excel.Workbooks.Open(admin_xls_fpath)
            admin_xls_converter_wb.SaveAs(f'{admin_xls_fpath}x', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            admin_xls_converter_wb.Close()
            excel.Application.Quit()

            # admin_xlsx 파일 불러오기
            admin_xlsx_wb = openpyxl.load_workbook(f'{admin_xls_fpath}x', data_only=True)

            # admin_xls 파일 첫번째 셀 열기
            admin_xlsx_ws = admin_xlsx_wb.active

            # mango_xlsx 파일 불러오기
            mango_xlsx_wb = openpyxl.load_workbook(mango_xlsx_fpath, data_only=True)

            # mango_xlsx 파일 활성화된 셀 열기
            mango_xlsx_ws = mango_xlsx_wb.active

            if mango_xlsx_ws.max_row != admin_xlsx_ws.max_row:
                self.status_label2.setText("두 파일의 행 개수가 일치하지 않습니다. 두 개의 파일이 같은 데이터 파일인지 확인해주세요.")
                return 0

            else :
                self.status_label2.setText("더망고 송장 파일 생성 중입니다.....")
                QApplication.processEvents()

                for i in range(2,mango_xlsx_ws.max_row+1) : 
                    for j in range(2,mango_xlsx_ws.max_row+1) : 
                        if mango_xlsx_ws.cell(i,2).value == admin_xlsx_ws.cell(j,5).value :
                            mango_xlsx_ws.cell(row = i, column = 31, value = admin_xlsx_ws.cell(j,10).value)
                            break
                    if mango_xlsx_ws.cell(i,2).value == admin_xlsx_ws.cell(j,5).value :
                        continue

            # mango_UPLOAD 엑셀 저장
            mango_xlsx_wb.save(mango_xlsx_fpath)

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
            mango_xlsx_fpath = mango_xlsx_fpath.replace('/', '\\')

            # 더망고.xlsx 파일에서 확장자명 제외하고 파일명만 추출.
            mango_date_file_name_list = mango_xlsx_fpath.split('.')
            mango_date_file_name = mango_date_file_name_list[0]

            # xlsx to xlsx convert : openpyxl로 저장한 엑셀은 더망고에서 읽어지지 않음. 그래서 Microsoft excel을 열어서 저장을 한번 해줌.
            excel_converter2 = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xlsx_converter_wb = excel_converter2.Workbooks.Open(mango_xlsx_fpath)
            mango_xlsx_converter_wb.SaveAs(f'{mango_date_file_name}_더망고_업로드용.xlsx', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xlsx_converter_wb.Close()
            excel_converter2.Application.Quit()

            self.status_label2.setText("더망고 송장 업로드 파일 생성이 완료되었습니다.")
#-------------------------------------라이센스 관련 함수--------------------------------------------------------------------------------------
    # 라이센스 입력 함수
    def input_license(self) :
        with open(f'excel_conv_license_key.txt', mode='w') as file :            # txt 파일 읽어오기
            license_key = self.decry_license_text_line_edit.text()
            file.write(license_key)
        self.license_status.setText('라이센스 입력이 완료되었습니다.')
        QApplication.processEvents()
        self.decryptography()
        return

    # 복호화 함수
    def decryptography(self) :      
        # ex_conv)ad:as:ss:aa:aa:aa,2022-07-17
        # ex_conv)open,2022-07-17
        # 라이센스 파일 존재 유무 확인.
        license_file_existence = os.path.isfile(f'excel_conv_license_key.txt')      # 라이센스 파일 존재 유무 확인.    존재할 경우 True, 비존재할경우 False 반환.

        if license_file_existence == True :                              # 라이센스 파일이 존재할 경우 실행.
            with open(f'excel_conv_license_key.txt', mode='r') as file :            # txt 파일 읽어오기
                license_key = file.read()
                self.decry_license_text_line_edit.setText(license_key)
        elif license_file_existence == False :                                       # 라이센스 파일이 존재하지 않을 경우 실행.
            self.license_status.setText('엑셀변환기능을 사용하려면 라이센스를 입력하세요.\n프로그램이 설치된 경로내에 라이센스키가 들어있는 메모장이 생성됩니다.')
            return

        key_code = "0rKTqtZ4GQJlecUE8zRfU1B-metwX61_2Iz6B66B5eo="
        key_code = key_code.encode('utf-8')
        fernet = Fernet(key_code)

        today = dt.datetime.today()

        # 복호화 성공 시 실행
        try :
            decry_text = fernet.decrypt(license_key.encode('utf-8'))           # license를 인코딩하여 바이너리형태로 만들고, 복호화 하여 decry_text에 저장
            decry_text = decry_text.decode('utf-8')                             # decry_text를 디코딩하여 str형태로 변형.
        
            excel_conv_license_data = decry_text.split(',')

            self.this_mac_status.setText(getmac.get_mac_address())
                
            period_VPN_decry_mac = excel_conv_license_data[0]
            period_VPN_decry_expiration = excel_conv_license_data[1]
            period_VPN_decry_expiration_date_list = period_VPN_decry_expiration.split('-')            # mapping_decry_expiration을 -로 나눠 연,월,일로 나눔.
            period_VPN_decry_expiration_year = int(period_VPN_decry_expiration_date_list[0])          # 연 정보가 입력된 str 데이터를 int로 변형.
            period_VPN_decry_expiration_monce = int(period_VPN_decry_expiration_date_list[1])         # 월 정보가 입력된 str 데이터를 int로 변형.
            period_VPN_decry_expiration_day = int(period_VPN_decry_expiration_date_list[2])         # 일 정보가 입력된 str 데이터를 int로 변형.
            period_VPN_decry_expiration_dt = dt.datetime(period_VPN_decry_expiration_year,period_VPN_decry_expiration_monce,period_VPN_decry_expiration_day)    # 날짜 계산을 위해 연,월,일 정보를 합하여 datetime 데이터로 변형함.

            period_VPN_decry_service_period = period_VPN_decry_expiration_dt - today             # [만기일 - 현재일 = 서비스 기간] 식을 만들어 줌.
            period_VPN_decry_service_period_days = str(int(period_VPN_decry_service_period.days)+1)

            # 복호화 성공 후 서비스 기한이 양수인 경우 True 반환
            if int(period_VPN_decry_service_period_days) >= 0 :
                if period_VPN_decry_mac == getmac.get_mac_address() or period_VPN_decry_mac == 'open': 
                    self.make_admin_file_btn.setEnabled(True)
                    self.make_mangofile_btn.setEnabled(True)
                    self.license_status.setText('허용된 라이센스')
                    self.status_label.setText('허용된 라이센스')
                    self.status_label2.setText('허용된 라이센스')
                
                elif period_VPN_decry_mac != getmac.get_mac_address() and period_VPN_decry_mac != 'open':
                    self.make_admin_file_btn.setEnabled(False)
                    self.make_mangofile_btn.setEnabled(False)
                    self.license_status.setText('허용되지 않은 맥주소')
                    self.status_label.setText('허용되지 않은 맥주소')
                    self.status_label2.setText('허용되지 않은 맥주소')

            # 복호화 성공 후 서비스 기한이 음수인 경우 종료일을 보여주기 위해 종료 기한 날짜 반환.
            elif int(period_VPN_decry_service_period_days) < 0 :
                self.make_admin_file_btn.setEnabled(False)
                self.make_mangofile_btn.setEnabled(False)
                self.license_status.setText('사용기한 초과')
                self.status_label.setText('사용기한 초과')
                self.status_label2.setText('사용기한 초과')

            # 라벨에 맥, 남은기간, 날짜 출력.
            if excel_conv_license_data[0] == 'open' :
                self.period_VPN_decry_mac_status.setText('ALL')

            elif excel_conv_license_data != 'open' :
                self.period_VPN_decry_mac_status.setText(period_VPN_decry_mac)

            self.period_VPN_decry_expiration_status.setText(period_VPN_decry_expiration)
            self.period_VPN_decry_service_period_status.setText(f'{period_VPN_decry_service_period_days}일')    # label에 서비스 기간 출력.

        # 복호화 실패 시 오류가 뜰 때 실행
        except :
            self.make_admin_file_btn.setEnabled(False)
            self.make_mangofile_btn.setEnabled(False)
            self.license_status.setText('허가되지않은 라이센스입니다.')
            self.status_label.setText('허가되지않은 라이센스입니다.')
            self.status_label2.setText('허가되지않은 라이센스입니다.')

            self.period_VPN_decry_mac_status.setText('사용불가')
            self.this_mac_status.setText(getmac.get_mac_address())
            self.period_VPN_decry_expiration_status.setText('사용불가')
            self.period_VPN_decry_service_period_status.setText('사용불가')

    # 스마트스토어 구매 링크
    def smartstore(self) :
        webbrowser.open('https://smartstore.naver.com/ctmall_/products/6926992498')
        return

    # 더망고 사용법 블로그 링크
    def excel_converter_manual_blog(self) :
        webbrowser.open('https://blog.naver.com/awldnjs2/222816206164')
        return


QApplication.setStyle("fusion")
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()
sys.exit(app.exec_())