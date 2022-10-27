from msilib.schema import RemoveFile
import win32com.client as win32
import xlrd
import openpyxl
from tkinter import filedialog
from PyQt5.QtWidgets import *
from PyQt5 import uic
import sys
import time
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UI_PATH = "Excel_converter_범용.ui"
info_file_fpath = '더망고 엑셀\info_file.xlsx'

class MainDialog(QDialog) :
    def __init__(self) :
        QDialog.__init__(self, None)
        uic.loadUi(os.path.join(BASE_DIR, UI_PATH), self)

#-------------------------------------CNP_파일_생성_버튼------------------------------------------------------------------------
        # 저장하기_버튼
        self.save_btn.clicked.connect(self.save_start)
        # 불러오기_버튼
        self.load_btn.clicked.connect(self.load_start)
        # 더망고_파일_선택_파일찾기_버튼
        self.search_mango_xls_btn.clicked.connect(self.search_mango_xls_start)
        # CNP_파일_생성_버튼
        self.make_CNPfile_btn.clicked.connect(self.make_CNPfile)
#-------------------------------------더망고_송장_파일_생성_버튼------------------------------------------------------------------------
        # CNP_파일_선택_파일찾기_버튼
        self.search_CNP_xls_btn.clicked.connect(self.search_invoice_CNP_xls_start)
        # 망고_파일_선택_파일찾기_버튼
        self.search_mango_xlsx_btn.clicked.connect(self.search_maked_mango_xlsx_start)
        # 더망고_송장_파일_생성_버튼
        self.make_mangofile_btn.clicked.connect(self.make_invoice_mangofile_btn)

#-------------------------------------CNP_파일_생성_버튼_함수------------------------------------------------------------------------
    # 저장하기_버튼_함수
    def save_start(self) : 
        box_amount = self.box_amount_edit.text()
        box_type = self.box_type_edit.text()
        sender_name = self.sender_name_edit.text()
        sender_address = self.sender_address_edit.text()
        sender_phone = self.sender_phone_edit.text()

        # 저장하기
        # CNP 엑셀 파일 생성
        info_file_wb = openpyxl.Workbook()
        info_file_ws = info_file_wb.active

        # 첫 행 양식 생성 방법
        info_file_ws.append(["박스수량","박스타입","보내는분 성명","보내는분 주소","보내는분 전화번호"])
        info_file_ws.append([box_amount, box_type, sender_name, sender_address, sender_phone])
        info_file_wb.save(info_file_fpath)

        self.status_label.setText("저장하기가 완료되었습니다.")
    # 불러오기_버튼_함수
    def load_start(self) :
        info_file_wb = openpyxl.load_workbook(info_file_fpath, data_only=True)
        info_file_ws = info_file_wb.active

        box_amount = info_file_ws.cell(2,1).value
        box_type = info_file_ws.cell(2,2).value
        sender_name = info_file_ws.cell(2,3).value
        sender_address = info_file_ws.cell(2,4).value
        sender_phone = info_file_ws.cell(2,5).value

        self.box_amount_edit.setText(box_amount)
        self.box_type_edit.setText(box_type)
        self.sender_name_edit.setText(sender_name)
        self.sender_address_edit.setText(sender_address)
        self.sender_phone_edit.setText(sender_phone)

        self.status_label.setText("불러오기가 완료되었습니다.")
    # 더망고_파일_선택_파일찾기_버튼_함수
    def search_mango_xls_start(self) :
        # 더망고 xls 파일 경로 선택
        mango_xls_fname = filedialog.askopenfilename(initialdir="/", title = "더망고에서 다운받은 xlsx파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.mango_xls_box.setText(mango_xls_fname)
    # CNP_파일_생성_버튼_함수
    def make_CNPfile(self) :
        # 파일에 입력한 문자열 가져오기
        mango_xls_fname = self.mango_xls_box.toPlainText()
        box_amount = self.box_amount_edit.text()
        box_type = self.box_type_edit.text()
        sender_name = self.sender_name_edit.text()
        sender_address = self.sender_address_edit.text()
        sender_phone = self.sender_phone_edit.text()

        # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
        mango_xls_fname = mango_xls_fname.replace('/', '\\')

        # 더망고 xls 파일을 선택하지 않았을 시 경고 메시지.
        if mango_xls_fname == "" or box_amount == "" or box_type == "" or sender_name == "" or sender_address == "" or sender_phone == "" :
            self.status_label.setText("CNP 파일 생성란의 모든 값을 입력해주세요.")
            return 0
        
        else :
            self.status_label.setText("CNP 파일 생성이 진행중입니다...")
            QApplication.processEvents()

            # xls to xlsx convert
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xls_wb = excel.Workbooks.Open(mango_xls_fname)
            mango_xls_wb.SaveAs(mango_xls_fname+"x", FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xls_wb.Close()
            excel.Application.Quit()

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 더망고 xlsx 파일 위치 문자열
            mango_fpath = rf'{mango_xls_fname}x'

            # 더망고 xlsx 파일 불러오기
            mango_wb = openpyxl.load_workbook(mango_fpath, data_only=True)

            # 더망고 xlsx 파일 활성화된 시트 선택
            mango_ws = mango_wb.active

            # CNP 엑셀 파일 생성
            cnp_wb = openpyxl.Workbook()
            cnp_ws = cnp_wb.active

            # 첫 행 양식 생성 방법
            cnp_ws.append(["받는분성명","받는분전화번호","받는분기타연락처","받는분우편번호","받는분주소(전체, 분할)","박스수량","박스타입","배송메세지1","보내는분 성명","보내는분 주소","보내는분 전화번호","품목명"])

            # 데이터 이전
            for row in mango_ws.iter_rows(min_row=2) :
                if row[10].value == None :
                    cnp_ws.append([row[8].value, row[12].value, None, row[11].value, row[9].value,                      box_amount, box_type, row[14].value, sender_name, sender_address, sender_phone, f'{row[16].value}, [결제수량 : {row[18].value}]'])
                else :
                    cnp_ws.append([row[8].value, row[12].value, None, row[11].value, f'{row[9].value} {row[10].value}', box_amount, box_type, row[14].value, sender_name, sender_address, sender_phone, f'{row[16].value}, [결제수량 : {row[18].value}]'])

            # CNP 엑셀 저장
            cnp_wb.save(f'{mango_fpath}_CNP_UPLOAD.xlsx')

            # 위치 자동 내려받기
            self.mango_xlsx_box.setText(f'{self.mango_xls_box.toPlainText()}x')
            
            self.status_label.setText("CNP 파일 생성이 완료되었습니다.")
#-------------------------------------더망고_송장_파일_생성_버튼_함수------------------------------------------------------------------------
    # CNP_파일_선택_파일찾기_버튼_함수
    def search_invoice_CNP_xls_start(self) : 
        # CNP 송장 xls 파일 경로 선택
        invoice_CNP_xls_fname = filedialog.askopenfilename(initialdir="/", title = "CN Plus에서 다운받은 송장 xls파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.CNP_xls_box.setText(invoice_CNP_xls_fname)
    # 망고_파일_선택_파일찾기_버튼_함수
    def search_maked_mango_xlsx_start(self) : 
        # 더망고 xlsx 생성 파일 경로 선택
        maked_mango_xlsx_fname = filedialog.askopenfilename(initialdir="/", title = "생성된 더망고 xlsx파일을 선택 해 주세요", filetypes = (("xlsx file","*.xlsx"),("all file","*.*")))
        self.mango_xlsx_box.setText(maked_mango_xlsx_fname)
    # 더망고_송장_파일_생성_버튼_함수
    def make_invoice_mangofile_btn(self) :
        # 파일에 입력한 문자열 가져오기
        cnp_xls_fpath = self.CNP_xls_box.toPlainText()
        mango_xlsx_fpath = self.mango_xlsx_box.toPlainText()

        if cnp_xls_fpath == "" or mango_xlsx_fpath == "" :
            self.status_label2.setText("더망고 송장 파일 생성란의 모든 값을 입력해주세요.")
            return 0

        else :
            # CNP_xls 파일 불러오기
            cnp_xls_wb = xlrd.open_workbook(cnp_xls_fpath)

            # CNP_xls 파일 첫번째 셀 열기
            cnp_xls_ws = cnp_xls_wb.sheet_by_index(0)

            # mango_xlsx 파일 불러오기
            mango_xlsx_wb = openpyxl.load_workbook(mango_xlsx_fpath, data_only=True)

            # mango_xlsx 파일 활성화된 셀 열기
            mango_xlsx_ws = mango_xlsx_wb.active

            if mango_xlsx_ws.max_row != cnp_xls_ws.nrows:
                self.status_label2.setText("두 파일의 행 개수가 일치하지 않습니다. 두 개의 파일이 같은 데이터 파일인지 확인해주세요.")
                return 0

            else :
                self.status_label2.setText("더망고 송장 파일 생성 중입니다.....")
                QApplication.processEvents()

                for i in range(1,mango_xlsx_ws.max_row) : 
                    invoice_value = cnp_xls_ws.cell_value(i,7)
                    invoice_value = invoice_value.replace('-', '')
                    mango_xlsx_ws.cell(row = i+1, column = 31, value = invoice_value)
                
            # mango_UPLOAD 엑셀 저장
            mango_xlsx_wb.save(mango_xlsx_fpath)

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
            mango_xlsx_fpath = mango_xlsx_fpath.replace('/', '\\')

            # xlsx to xlsx convert : openpyxl로 저장한 엑셀은 더망고에서 읽어지지 않음. 그래서 Microsoft excel을 열어서 저장을 한번 해줌.
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xls_wb = excel.Workbooks.Open(mango_xlsx_fpath)
            mango_xls_wb.SaveAs(mango_xlsx_fpath+"_mango_UPLOAD.xlsx", FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xls_wb.Close()
            excel.Application.Quit()

            self.status_label2.setText("더망고 송장 업로드 파일 생성이 완료되었습니다.")

QApplication.setStyle("fusion")
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()

sys.exit(app.exec_())