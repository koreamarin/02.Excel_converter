from msilib.schema import RemoveFile
from posixpath import split
import win32com.client as win32
import openpyxl
from tkinter import filedialog
from PyQt5.QtWidgets import *
from PyQt5 import uic
import sys
import time
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UI_PATH = "hard_admin_Excel_converter.ui"

class MainDialog(QDialog) :
    def __init__(self) :
        QDialog.__init__(self, None)
        uic.loadUi(os.path.join(BASE_DIR, UI_PATH), self)
#-------------------------------------이지어드민_파일_생성_버튼_관련함수------------------------------------------------------------------------
        # 더망고_파일_선택_파일찾기_버튼
        self.search_mango_xls_btn.clicked.connect(self.search_mango_xls_start)
        # 이지어드민_파일_생성_버튼
        self.make_admin_file_btn.clicked.connect(self.make_adminfile)
#-------------------------------------더망고_송장_파일_생성_버튼_관련함수------------------------------------------------------------------------
        # 이지어드민_파일_선택_파일찾기_버튼
        self.search_admin_xls_btn.clicked.connect(self.search_invoice_admin_xls_start)
        # 망고_파일_선택_파일찾기_버튼
        self.search_mango_xlsx_btn.clicked.connect(self.search_maked_mango_xlsx_start)
        # 더망고_송장_파일_생성_버튼
        self.make_mangofile_btn.clicked.connect(self.make_invoice_mangofile_btn)

#-------------------------------------이지어드민_파일_생성_버튼_관련함수------------------------------------------------------------------------
    # 더망고_파일_선택_파일찾기_버튼_함수
    def search_mango_xls_start(self) :
        # 더망고 xls 파일 경로 선택
        mango_xls_fname = filedialog.askopenfilename(initialdir="/", title = "더망고에서 다운받은 xlsx파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.mango_xls_box.setText(mango_xls_fname)
        
    # 이지어드민_파일_생성_버튼_함수
    def make_adminfile(self) :
        # 파일에 입력한 문자열 가져오기
        mango_xls_fname = self.mango_xls_box.toPlainText()

        # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
        mango_xls_fname = mango_xls_fname.replace('/', '\\')

        # 더망고 xls 파일을 선택하지 않았을 시 경고 메시지.
        if mango_xls_fname == "" :
            self.status_label.setText("더망고에서 다운받은 xls 파일을 선택해주세요.")
            return 0
        
        else :
            self.status_label.setText("easy admin 업로드용 파일 생성이 진행중입니다...")
            QApplication.processEvents()

            # xls to xlsx convert
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xls_wb = excel.Workbooks.Open(mango_xls_fname)
            mango_xls_wb.SaveAs(f'{mango_xls_fname}x', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xls_wb.Close()
            excel.Application.Quit()

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 파일 날짜명
            admin_date_file_name_list = mango_xls_fname.split('.')
            admin_date_file_name = admin_date_file_name_list[0]

            # 더망고 xlsx 파일 불러오기
            mango_wb = openpyxl.load_workbook(f'{mango_xls_fname}x', data_only=True)

            # 더망고 xlsx 파일 활성화된 시트 선택
            mango_ws = mango_wb.active

            # 이지어드민 엑셀 파일 생성
            admin_wb = openpyxl.Workbook()
            admin_ws = admin_wb.active


            # 첫 행 양식 생성 방법
            data = []
            for column in range(1, mango_ws.max_column + 1) :
                data.append(mango_ws.cell(1,column).value)
            admin_ws.append(data)

            # 데이터 이전
            for row in mango_ws.iter_rows(min_row=2) :
                if row[10].value == None :
                    admin_ws.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value,                        row[10].value, row[11].value, row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value, row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value, row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value, row[36].value])
                else :
                    admin_ws.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, f'{row[9].value} {row[10].value}', None,            row[11].value, row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value, row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value, row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value, row[36].value])

            # 이지어드민 엑셀 저장
            admin_wb.save(f'{admin_date_file_name}_변환.xlsx')

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # xls to xlsx convert
            excel2 = win32.gencache.EnsureDispatch('Excel.Application')
            admin_xlsx_wb = excel2.Workbooks.Open(f'{admin_date_file_name}_변환.xlsx')
            admin_xlsx_wb.SaveAs(f'{admin_date_file_name}_이지어드민_업로드용.xlsx', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            admin_xlsx_wb.Close()
            excel2.Application.Quit()

            # 위치 자동 내려받기
            self.mango_xlsx_box.setText(f'{mango_xls_fname}x')
            
            self.status_label.setText("이지어드민 업로드용 파일 생성이 완료되었습니다.")
#-------------------------------------더망고_송장_파일_생성_버튼_함수------------------------------------------------------------------------
    # 이지어드민_파일_선택_파일찾기_버튼_함수
    def search_invoice_admin_xls_start(self) : 
        # 이지어드민 송장 xls 파일 경로 선택
        invoice_admin_xls_fname = filedialog.askopenfilename(initialdir="/", title = "CN Plus에서 다운받은 송장 xls파일을 선택 해 주세요", filetypes = (("xls file","*.xls"),("all file","*.*")))
        self.admin_xls_box.setText(invoice_admin_xls_fname)
    # 더망고_파일_선택_파일찾기_버튼_함수
    def search_maked_mango_xlsx_start(self) : 
        # 더망고 xlsx 생성 파일 경로 선택
        maked_mango_xlsx_fname = filedialog.askopenfilename(initialdir="/", title = "생성된 더망고 xlsx파일을 선택 해 주세요", filetypes = (("xlsx file","*.xlsx"),("all file","*.*")))
        self.mango_xlsx_box.setText(maked_mango_xlsx_fname)
    # 더망고_송장_파일_생성_버튼_함수
    def make_invoice_mangofile_btn(self) :
        # 파일에 입력한 문자열 가져오기
        admin_xls_fpath = self.admin_xls_box.toPlainText()
        mango_xlsx_fpath = self.mango_xlsx_box.toPlainText()

        # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
        admin_xls_fpath = admin_xls_fpath.replace('/', '\\')

        if admin_xls_fpath == "" or mango_xlsx_fpath == "" :
            self.status_label2.setText("이지어드민, 더망고 파일 위치를 선택해주세요.")
            return 0

        else :
            # xls to xlsx convert      # xls 파일이 이상해서 읽어지지 않음. 그래서 xlsx 파일로 변환.
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            admin_xls_converter_wb = excel.Workbooks.Open(admin_xls_fpath)
            admin_xls_converter_wb.SaveAs(f'{admin_xls_fpath}x', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            admin_xls_converter_wb.Close()
            excel.Application.Quit()

            # admin_xlsx 파일 불러오기
            admin_xlsx_wb = openpyxl.load_workbook(f'{admin_xls_fpath}x', data_only=True)

            # admin_xls 파일 첫번째 셀 열기
            admin_xlsx_ws = admin_xlsx_wb.active

            # mango_xlsx 파일 불러오기
            mango_xlsx_wb = openpyxl.load_workbook(mango_xlsx_fpath, data_only=True)

            # mango_xlsx 파일 활성화된 셀 열기
            mango_xlsx_ws = mango_xlsx_wb.active

            if mango_xlsx_ws.max_row != admin_xlsx_ws.max_row:
                self.status_label2.setText("두 파일의 행 개수가 일치하지 않습니다. 두 개의 파일이 같은 데이터 파일인지 확인해주세요.")
                return 0

            else :
                self.status_label2.setText("더망고 송장 파일 생성 중입니다.....")
                QApplication.processEvents()

                for i in range(2,mango_xlsx_ws.max_row+1) : 
                    for j in range(2,mango_xlsx_ws.max_row+1) : 
                        if mango_xlsx_ws.cell(i,2).value == admin_xlsx_ws.cell(j,5).value :
                            mango_xlsx_ws.cell(row = i, column = 31, value = admin_xlsx_ws.cell(j,10).value)
                            break
                    if mango_xlsx_ws.cell(i,2).value == admin_xlsx_ws.cell(j,5).value :
                        continue

            # mango_UPLOAD 엑셀 저장
            mango_xlsx_wb.save(mango_xlsx_fpath)

            # 저장될때까지 잠깐 쉼
            time.sleep(2)

            # 더망고 xls 파일 인식할 수 있도록 /를 \로 변환.  /를 사용했을 때 win32com이 인식을 못함.
            mango_xlsx_fpath = mango_xlsx_fpath.replace('/', '\\')

            # 더망고.xlsx 파일에서 확장자명 제외하고 파일명만 추출.
            mango_date_file_name_list = mango_xlsx_fpath.split('.')
            mango_date_file_name = mango_date_file_name_list[0]

            # xlsx to xlsx convert : openpyxl로 저장한 엑셀은 더망고에서 읽어지지 않음. 그래서 Microsoft excel을 열어서 저장을 한번 해줌.
            excel_converter2 = win32.gencache.EnsureDispatch('Excel.Application')
            mango_xlsx_converter_wb = excel_converter2.Workbooks.Open(mango_xlsx_fpath)
            mango_xlsx_converter_wb.SaveAs(f'{mango_date_file_name}_더망고_업로드용.xlsx', FileFormat = 51)   # FileFormat = 51 is for .xlsx extension
            mango_xlsx_converter_wb.Close()
            excel_converter2.Application.Quit()

            self.status_label2.setText("더망고 송장 업로드 파일 생성이 완료되었습니다.")

QApplication.setStyle("fusion")
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()
sys.exit(app.exec_())